import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open("data/audit.json") as f:
    data = json.load(f)

wb = Workbook()
thin_border = Border(
    bottom=Side(style="thin", color="D9D9D9")
)
header_font = Font(name="Arial", bold=True, size=11)
header_fill = PatternFill("solid", fgColor="F2F2F2")
body_font = Font(name="Arial", size=10)
red_font = Font(name="Arial", size=10, color="DC2626")
amber_font = Font(name="Arial", size=10, color="CA8A04")
green_font = Font(name="Arial", size=10, color="16A34A")
title_font = Font(name="Arial", bold=True, size=14)
subtitle_font = Font(name="Arial", size=10, color="666666")

def style_header_row(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

def auto_width(ws, min_w=10, max_w=60):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)

# ── Summary Sheet ──
ws = wb.active
ws.title = "Summary"
ws.sheet_properties.tabColor = "2563EB"

ws["A1"] = "SEO Audit: vitalcitynyc.org"
ws["A1"].font = title_font
ws["A2"] = f"Scanned {data['scannedAt'][:10]} \u2014 {data['pagesAudited']} of ~800 pages"
ws["A2"].font = subtitle_font

r = 4
for label, val in [
    ("Average Score", f"{data['averageScore']}/100"),
    ("Pages Audited", data["pagesAudited"]),
    ("Total Errors", data["totalIssues"]),
    ("Total Warnings", data["totalWarnings"]),
]:
    ws.cell(r, 1, label).font = Font(name="Arial", bold=True, size=10)
    ws.cell(r, 2, val).font = body_font
    r += 1

r += 1
ws.cell(r, 1, "ERRORS").font = header_font
ws.cell(r, 2, "COUNT").font = header_font
style_header_row(ws, r, 2)
r += 1
for issue, count in data["commonIssues"]:
    ws.cell(r, 1, issue).font = red_font
    ws.cell(r, 2, count).font = red_font
    r += 1

r += 1
ws.cell(r, 1, "WARNINGS").font = header_font
ws.cell(r, 2, "COUNT").font = header_font
style_header_row(ws, r, 2)
r += 1
for warning, count in data["commonWarnings"]:
    ws.cell(r, 1, warning).font = amber_font
    ws.cell(r, 2, count).font = amber_font
    r += 1

ws.column_dimensions["A"].width = 45
ws.column_dimensions["B"].width = 12

# ── Page Results Sheet ──
ws2 = wb.create_sheet("Page Results")
ws2.sheet_properties.tabColor = "16A34A"

headers = ["Score", "URL", "Title", "Meta Description", "Errors", "Warnings"]
for c, h in enumerate(headers, 1):
    ws2.cell(1, c, h)
style_header_row(ws2, 1, len(headers))

for i, page in enumerate(data["pages"], 2):
    score = page["score"]
    score_font = green_font if score >= 90 else (amber_font if score >= 70 else red_font)
    ws2.cell(i, 1, score).font = score_font
    ws2.cell(i, 2, page["url"].replace("https://www.vitalcitynyc.org", "")).font = body_font
    ws2.cell(i, 3, page.get("info", {}).get("title", "")).font = body_font
    ws2.cell(i, 4, page.get("info", {}).get("description", "")).font = body_font
    issues_str = "; ".join(page.get("issues", []))
    warnings_str = "; ".join(page.get("warnings", []))
    ws2.cell(i, 5, issues_str).font = red_font if issues_str else body_font
    ws2.cell(i, 6, warnings_str).font = amber_font if warnings_str else body_font
    for c in range(1, 7):
        ws2.cell(i, c).border = thin_border

ws2.column_dimensions["A"].width = 8
ws2.column_dimensions["B"].width = 50
ws2.column_dimensions["C"].width = 40
ws2.column_dimensions["D"].width = 50
ws2.column_dimensions["E"].width = 45
ws2.column_dimensions["F"].width = 45
ws2.auto_filter.ref = f"A1:F{len(data['pages']) + 1}"

# ── Duplicate Titles Sheet ──
ws3 = wb.create_sheet("Duplicate Titles")
ws3.sheet_properties.tabColor = "DC2626"

headers3 = ["Title", "Count", "Pages"]
for c, h in enumerate(headers3, 1):
    ws3.cell(1, c, h)
style_header_row(ws3, 1, len(headers3))

r = 2
for title, urls in data["duplicateTitles"]:
    ws3.cell(r, 1, title).font = body_font
    ws3.cell(r, 2, len(urls)).font = body_font
    ws3.cell(r, 3, "\n".join(u.replace("https://www.vitalcitynyc.org", "") for u in urls)).font = body_font
    ws3.cell(r, 3).alignment = Alignment(wrap_text=True)
    r += 1

ws3.column_dimensions["A"].width = 30
ws3.column_dimensions["B"].width = 10
ws3.column_dimensions["C"].width = 50

output = "/Users/joshgreenman/Experiments/.claude/worktrees/crazy-ptolemy/vital-city-seo-audit/seo-audit-vitalcitynyc.xlsx"
wb.save(output)
print(f"Saved to {output}")
